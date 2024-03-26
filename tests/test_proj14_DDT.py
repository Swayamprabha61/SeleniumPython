import pytest
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By


# Read the EXCEL to get the data from the file

def get_test_data():
    workbook = load_workbook('testdata.xlsx')
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data

@pytest.fixture
def driver():
    driver = webdriver.Chrome()
    driver.get("http://app.vwo.com")
    driver.maximize_window()
    yield driver
    driver.quit()


@pytest.mark.parametrize("username  , password, result", get_test_data())
def test_vwo_login(driver,username, password,result):
    driver.find_element(By.ID, "login-username").send_keys(username)
    driver.find_element(By.ID, "login-password").send_keys(password)
    driver.find_element(By.ID, "js-login-btn").click()
    print(username, password, driver.current_url)
    if result == "fail":
        error_msg = driver.find_element(By.ID, "js-notification-box-msg").text
        assert error_msg in "Your email, password, IP address or location did not match"
    else:
        assert "login" in driver.current_url
